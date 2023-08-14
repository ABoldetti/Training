from openpyxl import load_workbook
from openpyxl.styles import Font 
from openpyxl.worksheet.table import Table

# wb = Workbook()
# w0 = wb.active
# w0.title = 'Trial0'
# w1 = wb.create_sheet( "Trial1")
# w2 = wb.create_sheet( "Trial2")
# w3 = wb.create_sheet( "Trial3")
# for sheet in wb :
#     print( sheet.title)

# # method to add something to the file

# tree = [ [ 1 , 2 , 3 ] , [ 4 , 5 , 6 ] , [ 7 , 8 , 9 ] ]
# for i in tree :
#     w0.append( i )

# ft = Font(bold=True)
# for i in w0[ 'A1:C1']:
#     for cell in i:
#         cell.font = ft
# wb.save( 'trial.xlsx' )

# method to open a file
wb = load_workbook( '/Users/andreaboldetti/Documents/GitHub/My_first_Repository/trial.xlsx' )
ws = wb.active
# pls don't touch this code, i don't know how it works. tables.values gives back a shit of a string
for table in ws.tables.values() :
    a = str(table)
    b = a.split("Parameters:")
    c = list()
    c = b[2]
    d = c.split(",")
    i = d[0]
    e = i[slice( 6 , len( i ) -1 ) ]
    string = e.split( ":")
    print( type(string) )
print((string[0]))
for i in range( len( string ) ):
    ausy = 0
    for j in range( len( string[i] ) ):
        if ord( string[i][j] ) < 65:
            ausy = j
            break
    n = string[ i ][ slice( j , len( string [ i ] ) ) ]
    char = string[ i ][ slice( j ) ]
    print( char , n , f"""{char}{n}""")
            

# print( a )
# for table in ws.tables:
#     print( table , "\n\n\n\n\n\n")
# a.split( " " )
# for i in a:
#     i.split('=')
#     print (i)
# for i in a:
#     if i == 'ref':
#         print( i, i+1)
#a=ws.tables.items
#print( type(a) )



# for i in range( len(b) ):
#     ausy = b[i].split(" ")
#     c.append(ausy)
#     print( b[i],'\n\n peppo \n\n' )

# for i in c:
#     if i.__contains__('ref='):
#         ausy = i
# #print( ausy )

# cell = ws['Nome cella']
# cell.value per ottenere il valore


