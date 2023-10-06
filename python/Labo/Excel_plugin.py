
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table
import pandas as pd

# class to take the data from an excel, to use it input the path of the excel file and use rolling_tables
# if the table is larger than 3 columns, it detect wether there is an 's' or an 'f' over the table to use them as 'start' and 'finish'
class excel :
    def __init__( self , path: str , table: bool) -> None:
        self.wb = load_workbook( path , data_only = True )
        self.ws = self.wb.active
        pass
    # function to call to start to get the values, if you wanna use it alone, use the for cicle and substitute n with table
    def rolling_table ( self , n: str) -> dict :
        # for table in self.ws.tables.values:
        coordinates =self.mod_table( self.getting_coordinates( self.table( n ) ) )
        return {"data": self.accumulating_data( coordinates ), "coordinates": self.elaborating_coordinates(coordinates) }
    

    # function that work on that mess of Dict.value of ws.tables.values() and get the coordinates
    def table ( self , data: str) -> list:

        # pls don't touch this code, i don't know how it works. tables.values gives back a shit of a string
        a = str( data )
        b = a.split( "Parameters:" )
        c = list( )
        c = b [ 2 ]
        d = c.split( "," )
        h = d [ 0 ]
        e = h [ slice ( 6 , len( h ) -1 ) ]
        return e.split ( ":" )


    # function that divides the coordinates into letteral and numerical part into a list of the format
    # [initial letter, initial number, last letter, last number]
    def getting_coordinates ( self , set: list) -> list :
        a = list()
        for i in range( len( set ) ):
            ausy = 0
            for j in range( len( set[i] ) ):
                if ord( set[i][j] ) < 65:
                    ausy = j
                    break
            n = int( set[i][ slice( j , len( set [ i ] ) ) ])
            char = set[ i ][ slice( j ) ]
            a.append( char )
            a.append( n )
        return a


    # function that gets the coordinates divided in list (First column, first row, last column, last row) and returns a DataFrame 
    # containing all the data
    def accumulating_data ( self , coordinates :list ) ->  pd.DataFrame:

        coordinates = self.checkingtable( coordinates )
        ausy = list()
        df = pd.DataFrame(ausy)

        for i in range( ord( coordinates [ 0 ] ) , ord( coordinates [ 2 ] ) + 1 ):
            ausy = list()
            for j in range( coordinates [ 1 ] + 1 , coordinates [ 3 ]+1 ):
                ausy.append( (self.ws[f"{chr(i)}{j}"].value) )
            df[ str(self.ws[f"{chr(i)}{coordinates[1]}"].value) ] = pd.DataFrame(ausy)
        return(df)
    
    # function used to detect the size of the table, if it's bigger than 3 columns it get reduced using mod_table
    def checkingtable( self , coord : list ) -> list :
        n_col = ord(coord[2]) - ord(coord[0])
        if n_col > 3: return self.mod_table( coord )
        else : return coord
    
    # function that uses the cells imediately over the table to try to reduce it to a table of len 3 or less
    def mod_table( self , coord ):
        for i in range ( ord( coord[0] ) , ord( coord[2] ) ):
            if self.ws[ f"{chr(i)}{coord[1]-1}"].value == 's' : coord[0] = chr( i )
            if self.ws[ f"{chr(i)}{coord[1]-1}"].value == 'f' : coord[2] = chr( i )
        if coord[0] < coord[2] : None
        else :
            ausy = coord[0]  
            coord[0] = coord[2]
            coord[2] = ausy
        return coord

    # function to elaborate the coordinates divided in a list after passing through "getting_coordinates"
    def elaborating_coordinates( self , coordinates: list ):
        return [ coordinates[ 0 ] , coordinates[ len(coordinates)-1 ] + 2 ]





if __name__ == '__main__' :
    a = excel( 'trial.xlsx' )
    print( a.rolling_table(1))